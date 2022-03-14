from openpyxl import workbook, load_workbook
import operator
from openpyxl.styles import PatternFill


wb = load_workbook(r"C:\Users\Administrator\Desktop\1.xlsx")
# 1、获取sheet名称
name_list = wb.sheetnames
print(name_list)

# for sheet_object in wb:
#     print(sheet_object.)
sheet = wb.worksheets[0]
sheet2 = wb.worksheets[1]
# print(sheet.rows)
row_data1 = []
all_data1 = []

row_data2 = []
all_data2 = []

row_title1 = []
row_title2 = []
# 读取标题
row_title1 = ["海典-" + cell.value for cell in sheet[1]]
row_title2 = ["ebs-" + cell.value for cell in sheet2[1]]
# 读取数据
for row in sheet.iter_rows(min_row=2):
    row_data1 = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in row
    ]
    if row_data1[20] == 0:
        row_data1[20] = None
    all_data1.append(row_data1)

for row in sheet2.iter_rows(min_row=2):
    row_data2 = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in row
    ]
    all_data2.append(row_data2)
# for row in sheet.rows:
#     print(row[0].value)
# 按行读取
# for row in sheet.iter_rows(min_row=1, max_row=3):

# 列表排序
all_data1.sort(key=lambda data1: data1[1])
all_data2.sort(key=lambda data2: data2[1])
# 排序后插入标题行
all_data1.insert(0, row_title1)
all_data2.insert(0, row_title2)

# 数据比对 删除重复项
# 从后往前删除 避免 list index out of range
for i in range(len(all_data1) - 1, 0, -1):
    if operator.eq(all_data1[i][:-1], all_data2[i][:-1]):
        del all_data1[i]
        del all_data2[i]

# 写入excel
wb2 = workbook.Workbook()
sheet3 = wb2.worksheets[0]

new_cols = len(row_data1) * 2

for row in range(1, len(all_data1) + 1):
    for col in range(1, new_cols + 1):
        if col % 2 == 0:
            sheet3.cell(row, col).value = all_data2[row - 1][int(col / 2) - 1]
        else:
            sheet3.cell(row, col).value = all_data1[row - 1][int(col / 2)]

orange_fill = PatternFill(fill_type="solid", fgColor="FFC125")

for i in range(2, len(all_data1) + 1):
    for j in range(1, sheet3.max_column, 2):
        if sheet3.cell(i, j).value != sheet3.cell(i, j + 1).value:
            sheet3.cell(i, j).fill = orange_fill
            sheet3.cell(i, j + 1).fill = orange_fill

wb2.save(r"C:\Users\Administrator\Desktop\河南新稀特前后端数据差异.xlsx")
