from openpyxl import workbook, load_workbook
import operator
from openpyxl.styles import PatternFill

wb = load_workbook(r"C:\Users\Administrator\Desktop\周口对比数据.xlsx")
# 1、获取sheet名称
name_list = wb.sheetnames
print(name_list)

sheet = wb.worksheets[0]
sheet2 = wb.worksheets[1]
print(sheet.max_row)
row_data1 = []
all_data1 = {}

row_data2 = []
all_data2 = {}

row_title1 = []
row_title2 = []
# 读取标题
row_title1 = ["批发-" + cell.value for cell in sheet[1]]
row_title2 = ["ebs-" + cell.value for cell in sheet2[1]]

# 读取数据到字典
for row in sheet.iter_rows(min_row=2):
    row_data1 = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in row
    ]
    all_data1[row_data1[0]] = row_data1

# print(all_data1)
for row in sheet2.iter_rows(min_row=2):
    row_data2 = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in row
    ]
    all_data2[row_data2[0]] = row_data2
# 1、去除没有相同键值的项
# 2、删除完全相同的项
for i in set(all_data1.keys()) | set(all_data2.keys()):
    if all_data1.get(i) is None:
        del [all_data2[i]]
        # print(i)
        continue
    if all_data2.get(i) is None:
        del [all_data1[i]]
        continue
    if operator.eq(all_data1[i], all_data2[i]):
        del [all_data1[i]]
        del [all_data2[i]]
# print(all_data1)
# print(len(all_data2))
print(all_data1)
# 写入excel
wb2 = workbook.Workbook()
sheet3 = wb2.worksheets[0]

new_cols = len(row_data1) * 2

row = 1
for i in all_data1.keys():
    for col in range(1, new_cols + 1):
        if col % 2 == 0:
            sheet3.cell(row, col).value = all_data2[i][int(col / 2) - 1]
        else:
            sheet3.cell(row, col).value = all_data1[i][int(col / 2)]
    row += 1
# 写入标题
sheet3.insert_rows(1)
for col in range(1, new_cols + 1):
    if col % 2 == 0:
        sheet3.cell(1, col).value = row_title2[int(col / 2) - 1]
    else:
        sheet3.cell(1, col).value = row_title1[int(col / 2)]

orange_fill = PatternFill(fill_type="solid", fgColor="FFC125")

for i in range(2, len(all_data1) + 2):
    for j in range(1, sheet3.max_column, 2):
        if sheet3.cell(i, j).value != sheet3.cell(i, j + 1).value:
            sheet3.cell(i, j).fill = orange_fill
            sheet3.cell(i, j + 1).fill = orange_fill

wb2.save(r"C:\Users\Administrator\Desktop\周口对比数据数据差异.xlsx")
