from openpyxl import load_workbook

# 读取excel数据
wb = load_workbook(filename=r"C:\Users\Administrator\Desktop\1.xlsx")
print(wb.sheetnames)
ws = wb.worksheets[0]
print(ws.values)
'''
遍历所有数据 1
for row in ws.values:
    for value in row:
        print(value)

遍历所有数据 2
for row in ws.iter_rows(values_only=True):
    print(row)
'''

