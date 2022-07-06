from openpyxl import load_workbook, workbook

wb = load_workbook(r"C:\Users\Administrator\Desktop\河南 编译后.xlsx")

sheet = wb.worksheets[0]
j = 2
for row in sheet.iter_rows(min_row=2):
    data_row = [cell.value for cell in row]
    if data_row[2] == "重新编译":
        if data_row[6] == "PACKAGE BODY":
            sql_exc = (
                "alter PACKAGE " + data_row[0] + "." + data_row[1] + " compile body;"
            )
        else:
            sql_exc = (
                "alter "
                + data_row[6]
                + " "
                + data_row[0]
                + "."
                + data_row[1]
                + " compile;"
            )
    else:
        sql_exc = "drop " + data_row[6] + " " + data_row[0] + "." + data_row[1] + ";"
    print(sql_exc)
    sheet.cell(j, 4).value = sql_exc
    j += 1

wb.save(r"C:\Users\Administrator\Desktop\河南 编译后.xlsx")
